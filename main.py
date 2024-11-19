import pandas as pd
import re
import os
import threading
import logging
from typing import List, Optional
import openpyxl
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Configure logging
logging.basicConfig(
    filename='email_generator_debug.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class FileHandler:
    @staticmethod
    def load_database(file_path: str, progress_callback: callable, cancel_event: threading.Event) -> Optional[pd.DataFrame]:
        logging.debug(f"Loading database from: {file_path}")
        file_extension = file_path.split('.')[-1].lower()

        if file_extension == 'csv':
            return FileHandler._load_csv(file_path, progress_callback, cancel_event)
        elif file_extension in ['xlsx', 'xls']:
            return FileHandler._load_excel(file_path, progress_callback, cancel_event)
        else:
            logging.error(f"Unsupported file format: {file_extension}")
            raise ValueError(f"Unsupported file format: {file_extension}")

    @staticmethod
    def _load_csv(file_path: str, progress_callback: callable, cancel_event: threading.Event) -> Optional[pd.DataFrame]:
        logging.debug("Loading CSV file")
        chunks = []
        chunk_size = 10000
        total_rows = sum(1 for _ in open(file_path, 'r')) - 1  # Subtract 1 for header
        
        for i, chunk in enumerate(pd.read_csv(file_path, chunksize=chunk_size, low_memory=False)):
            if cancel_event.is_set():
                logging.info("CSV loading cancelled")
                return None
            chunks.append(FileHandler.process_dataframe(chunk))
            if i % 10 == 0:  # Update progress less frequently
                progress = min(100, int((i * chunk_size / total_rows) * 100))
                progress_callback(progress)
        
        progress_callback(100)
        result = pd.concat(chunks, ignore_index=True)
        logging.debug(f"CSV loaded. Shape: {result.shape}")
        return result

    @staticmethod
    def _load_excel(file_path: str, progress_callback: callable, cancel_event: threading.Event) -> Optional[pd.DataFrame]:
        logging.debug("Loading Excel file")
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        all_data = []
        total_sheets = len(workbook.sheetnames)
        
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            if cancel_event.is_set():
                logging.info("Excel loading cancelled")
                return None
            sheet = workbook[sheet_name]
            data = list(sheet.values)
            df = pd.DataFrame(data[1:], columns=data[0])
            all_data.append(FileHandler.process_dataframe(df))
            if sheet_index % 2 == 0:  # Less frequent progress updates
                progress = min(100, int(((sheet_index + 1) / total_sheets) * 100))
                progress_callback(progress)
        
        progress_callback(100)
        result = pd.concat(all_data, ignore_index=True)
        logging.debug(f"Excel loaded. Shape: {result.shape}")
        return result

    @staticmethod
    def process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
        logging.debug("Processing dataframe")
        rename_dict = {
            'Company Name': 'company_name',
            'Domain': 'domain',
            'Email Format': 'format'
        }
        df.rename(columns=rename_dict, inplace=True)
        logging.debug(f"Processed dataframe. Columns: {df.columns}")
        return df

class EmailGenerator:
    titles_to_exclude = []  # Cache titles to exclude to avoid multiple file reads

    @staticmethod
    def load_exclusion_titles():
        if not EmailGenerator.titles_to_exclude:
            df = pd.read_csv('titles.csv')
            EmailGenerator.titles_to_exclude = df['Title'].tolist()

    @staticmethod
    def clean_suffixes(name: str) -> tuple:
        """Clean suffixes and titles from the name."""
        EmailGenerator.load_exclusion_titles()
        name_parts = name.split()
        cleaned_name_parts = [part for part in name_parts if part.lower() not in EmailGenerator.titles_to_exclude]
        removed_titles = [part for part in name_parts if part.lower() in EmailGenerator.titles_to_exclude]
        return ' '.join(cleaned_name_parts).strip(), removed_titles

    @staticmethod
    def process_name(first_name: str, last_name: str) -> tuple:
        """Process first and last names according to the new rules."""
        df = pd.read_csv('names_to_remove.csv')
        prefixes_to_remove = set(df['Title'].tolist())
        first_name_parts = first_name.split()
        first_name_parts = [part for part in first_name_parts if part.lower() not in prefixes_to_remove]

        # Move extra words from first name to last name
        if len(first_name_parts) > 1:
            last_name = ' '.join(first_name_parts[1:] + [last_name])
            first_name = first_name_parts[0]
        
        return first_name, last_name

    @staticmethod
    def get_domains(company_name: str, db_data: pd.DataFrame) -> List[str]:
        """Retrieve domains for a given company."""
        if pd.isna(company_name):
            return []

        matches = db_data[db_data['company_name'].str.contains(re.escape(company_name), case=False, na=False)]
        if matches.empty:
            return ["Not Recognized", "example.com"]

        return matches['domain'].unique().tolist()

    @staticmethod
    def get_email_formats(company_name: str, db_data: pd.DataFrame) -> List[str]:
        """Retrieve email formats for a given company."""
        if pd.isna(company_name):
            return []

        matches = db_data[db_data['company_name'].str.contains(re.escape(company_name), case=False, na=False)]
        return matches['format'].dropna().unique().tolist()

    @staticmethod
    def is_valid_email(email: str) -> bool:
        """Validate the format of an email address."""
        pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        return bool(re.match(pattern, email))

    @staticmethod
    def create_email_by_format(first_names: List[str], last_names: List[str], email_formats: List[str], domain: str) -> List[str]:
        """Generate emails based on format and domain."""
        emails = set()

        full_last_name = ' '.join(last_names)
        last_name_variations = {
            'combined': full_last_name.replace(' ', ''),
            'hyphenated': full_last_name.replace(' ', '-'),
            'split_dots': full_last_name.replace(' ', '.'),
            'parts': full_last_name.split()
        }

        for fn in first_names:
            fn = fn.lower()
            for email_format in email_formats:
                emails.update(EmailGenerator._apply_format(fn, last_name_variations, email_format, domain))

        return list(emails)

    @staticmethod
    def _apply_format(first_name: str, last_name_variations: dict, email_format: str, domain: str) -> List[str]:
        """Helper function to apply the email format and generate emails."""
        emails = []
        combined = last_name_variations['combined']
        format_mappings = {
            "firstname.lastname": lambda: [f"{first_name}.{combined}@{domain}"],
            "firstnamelastname": lambda: [f"{first_name}{combined}@{domain}"],
            "firstname": lambda: [f"{first_name}@{domain}"],
            "lastname": lambda: [f"{combined}@{domain}"]
        }
        
        if email_format in format_mappings:
            emails.extend(format_mappings[email_format]())
        
        return emails

class EmailProcessor:
    def __init__(self, db_data: pd.DataFrame, progress_callback: callable, cancel_event: threading.Event):
        self.db_data = db_data
        self.progress_callback = progress_callback
        self.cancel_event = cancel_event
        self.email_count = 0  # Counter for generated emails
        logging.debug("EmailProcessor initialized")

    def process_contact_file(self, file_path: str) -> None:
        logging.info(f"Processing contact file: {file_path}")
        try:
            data = FileHandler.load_database(file_path, self.progress_callback, self.cancel_event)
            
            if data is None:  # Loading was cancelled
                logging.info("Contact file loading cancelled")
                return

            logging.debug(f"Contact data loaded. Shape: {data.shape}")
            logging.debug(f"Contact data columns: {data.columns}")
            logging.debug(f"First few rows of contact data:\n{data.head()}")

            required_columns = ['first_name', 'last_name', 'company']
            missing_columns = [col for col in required_columns if col not in data.columns]
            if missing_columns:
                error_msg = f"Required columns {', '.join(missing_columns)} not found in the file."
                logging.error(error_msg)
                raise ValueError(error_msg)
            
            if 'Email' not in data.columns:
                data['Email'] = ''
            if 'domain' not in data.columns:
                data['domain'] = ''

            titles_to_exclude = EmailGenerator.titles_to_exclude

            total_rows = len(data)
            for index, row in data.iterrows():
                if self.cancel_event.is_set():
                    logging.info(f"Processing stopped. Emails generated: {self.email_count}")
                    break
                self._process_row(row, data, titles_to_exclude)
                if index % 10 == 0:  # Progress updated less frequently
                    self.progress_callback(min(100, int((index + 1) / total_rows * 100)))

            self._save_output(file_path, data)
            logging.info(f"Email generation completed. Total emails generated: {self.email_count}")
            messagebox.showinfo("Success", f"Emails generated successfully! Total: {self.email_count}")
        except Exception as e:
            logging.error(f"Error processing contact file: {e}", exc_info=True)
            messagebox.showerror("Error", str(e))

    def _process_row(self, row: pd.Series, data: pd.DataFrame, titles_to_exclude: List[str]) -> None:
        logging.debug(f"Processing row: {row}")
        first_name = row['first_name'].strip() if pd.notna(row['first_name']) else ""
        last_name = row['last_name'].strip() if pd.notna(row['last_name']) else ""
        company_name = row['company']

        if not first_name and not last_name:
            logging.warning(f"Skipping row due to missing both first and last names for company: {company_name}")
            return

        # Process the name
        first_name, last_name = EmailGenerator.process_name(first_name, last_name)

        # Clean the names and collect suffixes
        first_name_cleaned, suffixes_first = EmailGenerator.clean_suffixes(first_name)
        last_name_cleaned, suffixes_last = EmailGenerator.clean_suffixes(last_name)

        logging.debug(f"Processing: {first_name_cleaned} {last_name_cleaned}, Company: {company_name}")

        domains = EmailGenerator.get_domains(company_name, self.db_data)
        if not domains:
            logging.warning(f"No domain found for company: {company_name}")
            return

        domain = domains[0]
        if domain == "Not Recognized":
            domain = "example.com"

        email_formats = EmailGenerator.get_email_formats(company_name, self.db_data)
        
        emails = EmailGenerator.create_email_by_format([first_name_cleaned], [last_name_cleaned], email_formats, domain)
        valid_emails = [email.lower() for email in emails if EmailGenerator.is_valid_email(email)]
        
        logging.debug(f"Valid emails generated: {valid_emails}")

        if valid_emails:
            data.at[row.name, 'Email'] = valid_emails[0]  # Use the first valid email
            data.at[row.name, 'domain'] = domain
            data.at[row.name, 'first_name'] = first_name_cleaned
            data.at[row.name, 'last_name'] = last_name_cleaned
            self.email_count += 1
        else:
            logging.warning(f"No valid emails generated for {first_name_cleaned} {last_name_cleaned} at {company_name}")

    def _save_output(self, file_path: str, data: pd.DataFrame) -> None:
        # Save updated data back to the original file
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension == '.csv':
            data.to_csv(file_path, index=False)
        elif file_extension in ['.xlsx', '.xls']:
            data.to_excel(file_path, index=False)
        
        logging.info(f"Updated data saved to {file_path}")
        logging.info(f"Number of emails generated: {self.email_count}")

class EmailGeneratorApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Email Generator")
        self.master.geometry("400x250")

        self.progress_var = tk.IntVar()
        self.cancel_event = threading.Event()

        self.load_database_button = ttk.Button(master, text="Load Database", command=self.load_database)
        self.load_database_button.pack(pady=10)

        self.load_contacts_button = ttk.Button(master, text="Load Contacts", command=self.load_contacts)
        self.load_contacts_button.pack(pady=10)

        self.progress_bar = ttk.Progressbar(master, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(pady=10)

        self.cancel_button = ttk.Button(master, text="Cancel", command=self.cancel_operation, state=tk.DISABLED)
        self.cancel_button.pack(pady=10)

        self.status_label = ttk.Label(master, text="")
        self.status_label.pack(pady=10)

        logging.debug("EmailGeneratorApp initialized")

    def load_database(self):
        logging.info("Loading database")
        file_path = filedialog.askopenfilename(title="Select Database File", filetypes=[("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv")])
        if file_path:
            self.cancel_event.clear()
            self.cancel_button.config(state=tk.NORMAL)
            self.load_database_button.config(state=tk.DISABLED)
            self.load_contacts_button.config(state=tk.DISABLED)
            
            def load_thread():
                try:
                    self.db_data = FileHandler.load_database(file_path, self.update_progress, self.cancel_event)
                    if self.db_data is not None:
                        logging.info("Database loaded successfully")
                        self.master.after(0, lambda: messagebox.showinfo("Success", "Database loaded successfully!"))
                    else:
                        logging.info("Database loading was cancelled")
                        self.master.after(0, lambda: messagebox.showinfo("Cancelled", "Database loading was cancelled."))
                except Exception as e:
                    logging.error(f"Error loading database: {e}", exc_info=True)
                    self.master.after(0, lambda: messagebox.showerror("Error", str(e)))
                finally:
                    self.master.after(0, self.reset_ui)

            threading.Thread(target=load_thread, daemon=True).start()

    def load_contacts(self):
        logging.info("Loading contacts")
        if not hasattr(self, 'db_data'):
            logging.warning("Database not loaded before attempting to load contacts")
            messagebox.showwarning("Warning", "Please load the database file first.")
            return

        file_path = filedialog.askopenfilename(title="Select Contacts File", filetypes=[("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv")])
        if file_path:
            self.cancel_event.clear()
            self.cancel_button.config(state=tk.NORMAL)
            self.load_database_button.config(state=tk.DISABLED)
            self.load_contacts_button.config(state=tk.DISABLED)

            def process_thread():
                try:
                    email_processor = EmailProcessor(self.db_data, self.update_progress, self.cancel_event)
                    email_processor.process_contact_file(file_path)
                except Exception as e:
                    logging.error(f"Error processing contacts: {e}", exc_info=True)
                    self.master.after(0, lambda: messagebox.showerror("Error", str(e)))
                finally:
                    self.master.after(0, self.reset_ui)

            threading.Thread(target=process_thread, daemon=True).start()

    def update_progress(self, value):
        self.progress_var.set(value)
        self.status_label.config(text=f"Processing... {value}%")
        logging.debug(f"Progress updated: {value}%")

    def cancel_operation(self):
        logging.info("Cancel operation requested")
        self.cancel_event.set()

    def reset_ui(self):
        logging.debug("Resetting UI")
        self.cancel_button.config(state=tk.DISABLED)
        self.load_database_button.config(state=tk.NORMAL)
        self.load_contacts_button.config(state=tk.NORMAL)
        self.progress_var.set(0)
        self.status_label.config(text="")

if __name__ == "__main__":
    logging.info("Starting Email Generator application")
    root = tk.Tk()
    app = EmailGeneratorApp(root)
    root.mainloop()
    logging.info("Email Generator application closed")
